from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_DIR = Path("outputs/training")

SERVICE_UNITS = [
    ("09940", "โรงพยาบาลส่งเสริมสุขภาพตำบลถนน"),
    ("09941", "โรงพยาบาลส่งเสริมสุขภาพตำบลตรัง"),
    ("09942", "โรงพยาบาลส่งเสริมสุขภาพตำบลกระหวะ"),
    ("09943", "โรงพยาบาลส่งเสริมสุขภาพตำบลลุโบะยิไร"),
    ("09944", "โรงพยาบาลส่งเสริมสุขภาพตำบลละงา"),
    ("09945", "โรงพยาบาลส่งเสริมสุขภาพตำบลกาเยาะมาตี"),
    ("09946", "โรงพยาบาลส่งเสริมสุขภาพตำบลเกาะจัน"),
    ("09947", "โรงพยาบาลส่งเสริมสุขภาพตำบลปะโด"),
    ("09948", "โรงพยาบาลส่งเสริมสุขภาพตำบลสาคอบน"),
    ("09949", "โรงพยาบาลส่งเสริมสุขภาพตำบลสาคอใต้"),
    ("09950", "โรงพยาบาลส่งเสริมสุขภาพตำบลสะกำ"),
    ("09951", "โรงพยาบาลส่งเสริมสุขภาพตำบลปานัน"),
    ("41083", "โรงพยาบาลส่งเสริมสุขภาพตำบลบ้านน้ำใส"),
    ("77483", "ศูนย์สุขภาพชุมชนตำบลมายอ"),
]

BASELINE_HEADERS = [
    "cid",
    "service_unit_code",
    "first_name",
    "last_name",
    "sex",
    "birth_date",
    "house_number",
    "village_no",
    "registry_status",
    "baseline_vaccine_status",
    "next_vaccine_due_date",
    "entry_type",
    "indicator_start_month",
    "is_ppa_target",
    "is_alternative_vaccine_target",
    "primary_vhv_name",
    "primary_family_health_volunteer_name",
]

LEGACY_HEADERS = [
    "#",
    "หน่วยบริการ",
    "ทะเบียนบุคคลPID",
    "เลขที่บัตรประชาชนCID",
    "เพศSEX",
    "ชื่อNAME",
    "นามสกุลLNAME",
    "อายุ(ปี)",
    "วันเกิดBIRTH",
    "บ้านเลขที่ ",
    "หมู่ ",
    "สถานะ วัคซีน ณ.เดือน พ.ค. 2569",
    "การ ศึกษา",
    "ชื่อสถานศึกษา",
    "สังกัด",
    "เป็นเป้าหมาย PPA",
    "เป็นเป้าหมาย รับวัคซีนทางเลือก",
    "ชื่อ อสม.ที่รับผิดชอบ",
    "ชื่อ ผรส.ที่รับผิดชอบ",
]

FIELD_NOTES = [
    ("cid", "เลขที่บัตรประชาชนCID", "เลข 13 หลัก ห้ามซ้ำ"),
    ("service_unit_code", "เลือกจากตั้งค่าหน่วยบริการ", "รหัส 5 หลักของหน่วยตน"),
    ("first_name", "ชื่อNAME", "ตัดช่องว่างหัวท้าย"),
    ("last_name", "นามสกุลLNAME", "ตัดช่องว่างหัวท้าย"),
    ("sex", "เพศSEX", "ชาย หรือ หญิง"),
    ("birth_date", "วันเกิดBIRTH", "รูปแบบ YYYY-MM-DD ตามไฟล์เดิม"),
    ("house_number", "บ้านเลขที่", "ต้องไม่ว่าง"),
    ("village_no", "หมู่", "ต้องไม่ว่าง"),
    ("registry_status", "ระบบเติมให้", "อยู่ในทะเบียนติดตาม"),
    ("baseline_vaccine_status", "สถานะ วัคซีน ณ.เดือน พ.ค. 2569", "ล่าช้าจะถูกแปลงเป็นค่ามาตรฐาน"),
    ("next_vaccine_due_date", "เติมเพิ่มเมื่อมีนัดหมายถัดไป", "เว้นว่างได้ ยกเว้นกรณีได้รับบางส่วนแต่ยังไม่ครบ"),
    ("entry_type", "ระบบเติมให้", "ทะเบียนตั้งต้น"),
    ("indicator_start_month", "ระบบเติมให้", "2026-06"),
    ("is_ppa_target", "เป็นเป้าหมาย PPA", "ใช่ / ไม่ใช่"),
    ("is_alternative_vaccine_target", "เป็นเป้าหมาย รับวัคซีนทางเลือก", "ใช่ / ไม่ใช่"),
    ("primary_vhv_name", "ชื่อ อสม.ที่รับผิดชอบ", "ต้องเติมชื่อผู้รับผิดชอบหลัก"),
    ("primary_family_health_volunteer_name", "ชื่อ ผรส.ที่รับผิดชอบ", "ต้องเติมชื่อผู้รับผิดชอบหลัก"),
]

STATUS_LIST = [
    "ล่าช้า-ยังไม่ได้เริ่ม/ไม่มีความคืบหน้า",
    "ได้รับบางส่วน-ยังไม่ครบเกณฑ์",
    "ได้รับบางส่วน-ล่าช้าต่อเนื่อง",
    "ปฏิเสธการฉีด",
    "ฉีดตามเกณฑ์",
]

HEADER_FILL = PatternFill("solid", fgColor="0F766E")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row=1, max_col=None):
    max_col = max_col or ws.max_column
    for cell in ws[row][:max_col]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def add_dropdown(ws, column, values):
    dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{column}2:{column}601")


def add_template_validations(ws):
    add_dropdown(ws, "E", ["ชาย", "หญิง"])
    add_dropdown(ws, "I", ["อยู่ในทะเบียนติดตาม"])
    add_dropdown(ws, "J", STATUS_LIST)
    add_dropdown(ws, "L", ["ทะเบียนตั้งต้น"])
    add_dropdown(ws, "N", ["ใช่", "ไม่ใช่"])
    add_dropdown(ws, "O", ["ใช่", "ไม่ใช่"])


def copy_row_formulas(formulas, target_row):
    return [formula.replace("2", str(target_row)) if formula else "" for formula in formulas]


def build_workbook(default_code=None, default_name=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "วิธีใช้"
    ws["A1"] = "ชุดไฟล์เตรียมทะเบียนตั้งต้นวัคซีนเด็ก อำเภอมายอ"
    ws["A1"].font = Font(size=18, bold=True, color="0F766E")
    instructions = [
        "1) เลือกหรือยืนยันรหัสหน่วยบริการในแท็บ ตั้งค่าหน่วยบริการ",
        "2) นำไฟล์ Excel/CSV เดิมของหน่วยมา copy เฉพาะข้อมูลเด็ก แล้ว paste ลงแท็บ ข้อมูลจากไฟล์เดิม ตั้งแต่แถวที่ 2",
        "3) เปิดแท็บ BASELINE_TEMPLATE ระบบจะดึง/แปลงข้อมูลหลักให้ตามหัวตารางมาตรฐาน 17 คอลัมน์",
        "4) เติมช่องที่ไฟล์เดิมมักไม่มี: primary_vhv_name และ primary_family_health_volunteer_name",
        "5) ตรวจแท็บ รายการที่ต้องตรวจ ก่อนส่งให้อำเภอหรือก่อน copy ไป Google Sheet กลาง",
        "6) เมื่อพร้อม ให้ copy ค่าใน BASELINE_TEMPLATE ไปวางใน Google Sheet กลาง หรือบันทึกเป็น TSV/CSV ตามที่แอดมินอำเภอกำหนด",
    ]
    for row_index, text in enumerate(instructions, start=3):
        ws[f"A{row_index}"] = text
        ws[f"A{row_index}"].alignment = Alignment(wrap_text=True)
    ws["A11"] = "ข้อสำคัญ: ห้ามแก้ชื่อหัวตารางภาษาอังกฤษใน BASELINE_TEMPLATE เพราะระบบอ่านด้วยชื่อหัวตารางนี้เท่านั้น"
    ws["A11"].fill = WARN_FILL
    ws["A11"].font = BOLD
    ws.column_dimensions["A"].width = 112

    cfg = wb.create_sheet("ตั้งค่าหน่วยบริการ")
    cfg.append(["service_unit_code", "service_unit_name", "active"])
    for code, name in SERVICE_UNITS:
        cfg.append([code, name, True])
    cfg["E1"] = "รหัสหน่วยบริการที่ใช้ในไฟล์นี้"
    cfg["F1"] = default_code or ""
    cfg["E2"] = "ชื่อหน่วยบริการ"
    cfg["F2"] = '=IFERROR(VLOOKUP(F1,A:B,2,FALSE),"")'
    cfg["E4"] = "หมายเหตุ"
    cfg["F4"] = "ถ้าเป็นไฟล์กลาง ให้เลือก/กรอกรหัสหน่วยบริการของหน่วยก่อนใช้งาน"
    style_header(cfg, 1, 3)
    for cell in ["E1", "E2", "E4"]:
        cfg[cell].font = BOLD
    cfg.column_dimensions["A"].width = 18
    cfg.column_dimensions["B"].width = 58
    cfg.column_dimensions["E"].width = 34
    cfg.column_dimensions["F"].width = 64
    dv = DataValidation(type="list", formula1=f"=$A$2:$A${len(SERVICE_UNITS) + 1}", allow_blank=False)
    cfg.add_data_validation(dv)
    dv.add("F1")

    old = wb.create_sheet("ข้อมูลจากไฟล์เดิม")
    old.append(LEGACY_HEADERS)
    style_header(old, 1, len(LEGACY_HEADERS))
    old.freeze_panes = "A2"
    for idx, header in enumerate(LEGACY_HEADERS, 1):
        old.column_dimensions[get_column_letter(idx)].width = max(14, min(32, len(header) + 4))
    old["A1"].comment = Comment("วางข้อมูลจากไฟล์เดิมให้ตรงหัวตารางนี้ เริ่มแถวที่ 2", "Codex")

    tpl = wb.create_sheet("BASELINE_TEMPLATE")
    tpl.append(BASELINE_HEADERS)
    style_header(tpl, 1, len(BASELINE_HEADERS))
    tpl.freeze_panes = "A2"
    formulas = [
        '=IF(\'ข้อมูลจากไฟล์เดิม\'!D2="","",TRIM(\'ข้อมูลจากไฟล์เดิม\'!D2))',
        '=IF(A2="","",\'ตั้งค่าหน่วยบริการ\'!$F$1)',
        '=IF(A2="","",TRIM(\'ข้อมูลจากไฟล์เดิม\'!F2))',
        '=IF(A2="","",TRIM(\'ข้อมูลจากไฟล์เดิม\'!G2))',
        '=IF(A2="","",TRIM(\'ข้อมูลจากไฟล์เดิม\'!E2))',
        '=IF(A2="","",\'ข้อมูลจากไฟล์เดิม\'!I2)',
        '=IF(A2="","",TRIM(\'ข้อมูลจากไฟล์เดิม\'!J2))',
        '=IF(A2="","",TRIM(\'ข้อมูลจากไฟล์เดิม\'!K2))',
        '=IF(A2="","","อยู่ในทะเบียนติดตาม")',
        '=IF(A2="","",IF(TRIM(\'ข้อมูลจากไฟล์เดิม\'!L2)="ล่าช้า","ล่าช้า-ยังไม่ได้เริ่ม/ไม่มีความคืบหน้า",TRIM(\'ข้อมูลจากไฟล์เดิม\'!L2)))',
        "",
        '=IF(A2="","","ทะเบียนตั้งต้น")',
        '=IF(A2="","","2026-06")',
        '=IF(A2="","",IF(TRIM(\'ข้อมูลจากไฟล์เดิม\'!P2)="เป็นเป้าหมาย PPA","ใช่","ไม่ใช่"))',
        '=IF(A2="","",IF(TRIM(\'ข้อมูลจากไฟล์เดิม\'!Q2)="เป็นเป้าหมาย รับวัคซีนทางเลือก","ใช่","ไม่ใช่"))',
        '=IF(A2="","",TRIM(\'ข้อมูลจากไฟล์เดิม\'!R2))',
        '=IF(A2="","",TRIM(\'ข้อมูลจากไฟล์เดิม\'!S2))',
    ]
    tpl.append(formulas)
    for row in range(3, 602):
        tpl.append(copy_row_formulas(formulas, row))
    add_template_validations(tpl)
    widths = [18, 18, 20, 20, 10, 14, 14, 10, 24, 42, 20, 18, 18, 16, 28, 26, 36]
    for idx, width in enumerate(widths, 1):
        tpl.column_dimensions[get_column_letter(idx)].width = width
    for col in ["P", "Q"]:
        for cell in tpl[f"{col}2:{col}601"]:
            cell[0].fill = WARN_FILL
    tpl["P1"].comment = Comment("ต้องเติมชื่อ อสม.หลัก ถ้าไฟล์เดิมไม่มี ระบบจะ reject", "Codex")
    tpl["Q1"].comment = Comment("ต้องเติมชื่อ ผรส.หลัก ถ้าไฟล์เดิมไม่มี ระบบจะ reject", "Codex")

    notes = wb.create_sheet("หัวตารางและคำอธิบาย")
    notes.append(["หัวตารางมาตรฐาน", "มาจากคอลัมน์ในไฟล์เดิม", "คำอธิบาย/กติกา"])
    for row in FIELD_NOTES:
        notes.append(row)
    style_header(notes, 1, 3)
    notes.column_dimensions["A"].width = 38
    notes.column_dimensions["B"].width = 44
    notes.column_dimensions["C"].width = 72
    for row in notes.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER

    checks = wb.create_sheet("รายการที่ต้องตรวจ")
    checks.append(["รายการตรวจ", "สูตร/วิธีดู", "ผ่านเมื่อ"])
    check_rows = [
        ["จำนวน CID ที่มีข้อมูล", "=COUNTA(BASELINE_TEMPLATE!A2:A601)", "ตรงกับจำนวนเด็กที่ต้องนำเข้า"],
        ["CID ไม่ครบ 13 หลัก", '=SUMPRODUCT(--(BASELINE_TEMPLATE!A2:A601<>""),--(LEN(BASELINE_TEMPLATE!A2:A601)<>13))', "ต้องเป็น 0"],
        ["ยังไม่เติม อสม.หลัก", '=COUNTIFS(BASELINE_TEMPLATE!A2:A601,"<>",BASELINE_TEMPLATE!P2:P601,"")', "ต้องเป็น 0 ก่อนส่ง"],
        ["ยังไม่เติม ผรส.หลัก", '=COUNTIFS(BASELINE_TEMPLATE!A2:A601,"<>",BASELINE_TEMPLATE!Q2:Q601,"")', "ต้องเป็น 0 ก่อนส่ง"],
        ["บ้านเลขที่ว่าง", '=COUNTIFS(BASELINE_TEMPLATE!A2:A601,"<>",BASELINE_TEMPLATE!G2:G601,"")', "ต้องเป็น 0"],
        ["หมู่ว่าง", '=COUNTIFS(BASELINE_TEMPLATE!A2:A601,"<>",BASELINE_TEMPLATE!H2:H601,"")', "ต้องเป็น 0"],
    ]
    for row in check_rows:
        checks.append(row)
    style_header(checks, 1, 3)
    checks.column_dimensions["A"].width = 34
    checks.column_dimensions["B"].width = 74
    checks.column_dimensions["C"].width = 30
    for row in checks.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = True
    return wb


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    master = build_workbook()
    master.save(OUTPUT_DIR / "baseline_import_template_all_units.xlsx")

    count = 0
    for code, name in SERVICE_UNITS:
        if code == "09941":
            continue
        wb = build_workbook(code, name)
        wb.save(OUTPUT_DIR / f"baseline_import_template_{code}.xlsx")
        count += 1

    print(f"created master and {count} unit files in {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
